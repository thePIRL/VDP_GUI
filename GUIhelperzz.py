import sys
import os
from random import shuffle
import numpy as np
from PIL import Image, ImageTk
import nibabel as nib
from scipy.signal import medfilt2d
import openpyxl

#------------------------------------------------------------------------------------
# -------------------------- HPG CLASS DEFINITION --------------------------------- ##
#------------------------------------------------------------------------------------
class HPG:
    """This may be new if you haven't done object oriented programming in the past,
but essentially I'm defining a new type of variable called HPG. The HPG variable
contains the HP and mask arrays which we give when we call the class, but there
are other attributes we might like to have such as the HP array normalized to 95th
percentile (normHP), and a defect array given a certain threshold (defectArray). What's
nice about classes is we can perform all of these calculations when we first create an
HPG variable - so they're calculated upon variable assignment and don't need to be
seprately run in the script."""
    def __init__(self,HP,mask,N4HP):
        rows,cols,slices = self.cropToData(mask)
        self.HP = HP[np.ix_(rows,cols,slices)]
        self.mask = mask[np.ix_(rows,cols,slices)]
        self.N4HP = N4HP[np.ix_(rows,cols,slices)]
        self.normHP  = self.normalizeHP() # -- always performed on N4 data
        self.HPmontage = self.HPtoMontage()
        self.calculateDefectArray(0.6)

    def normalizeHP(self):
        '''The N4 bias corrected data are normalized to their mean signal value for analysis'''
        normHP = np.divide(self.N4HP,np.mean(self.N4HP[self.mask>0]))
        return normHP
    
    def calculateDefectArray(self,thresh):
        '''given a threshold (specified as a fraction of the whole-lung signal mean) create the 3D binary defect array.
            Two arrays are created: the defect Array, and the border of the defect array (maybe helpful, idk...)'''
        defectArray = np.zeros(self.normHP.shape)
        for k in range(self.mask.shape[2]):
            defectArray[:,:,k] = medfilt2d((self.normHP[:,:,k]<thresh)*self.mask[:,:,k])
        maskBorder = np.zeros(self.normHP.shape)
        for k in range(self.mask.shape[2]):
            x = np.gradient(self.mask[:,:,k].astype(float))
            maskBorder[:,:,k] = (x[0]!=0)+(x[1]!=0)
        #self.defectBorder = defectBorder
        #self.defectArray = defectArray
        self.VDP = np.sum(defectArray)/np.sum(self.mask)*100
        self.defectArray = defectArray
        self.maskBorder = maskBorder

    def normalize95th(self):
        '''Normalize the HP array to its 95th percentile value. This is only for display in the GUI - no analysis is performed on these'''
        voxlist = self.HP[self.mask>0]
        voxlist.sort()
        norm95HP = np.divide(self.HP,voxlist[int(0.95*len(voxlist))])
        norm95HP[norm95HP>1] = 1
        norm95HP = norm95HP*255
        norm95HP = norm95HP.astype(np.uint8)
        return norm95HP

    def HPtoMontage(self,useBias=True):
        '''inputs 3D HP and returns 8bit 2D montage normalized to 95th for display'''
        if useBias:
            voxlist = self.N4HP[self.mask>0]
            voxlist.sort()
            HPimage = np.divide(self.N4HP,voxlist[int(0.99*len(voxlist))])
        else:
            voxlist = self.HP[self.mask>0]
            voxlist.sort()
            HPimage = np.divide(self.HP,voxlist[int(0.99*len(voxlist))])
        HPimage[HPimage>1] = 1
        HPimage = HPimage*255
        HPimage = HPimage.astype(np.uint8)
        l = [list(x) for x in HPimage.transpose(1,2,0).swapaxes(0,2)]
        self.HPmontage = np.block(l)
    
    def defectMontage(self,border=False):
        '''creates a 3D RGB array of the montaged data and mask for display'''
        combined3D = np.zeros((self.HPmontage.shape[0],self.HPmontage.shape[1],3))
        defectMontage = self.montage(self.defectArray)
        maskMontage = self.montage(self.maskBorder)
        if border:
            combined3D[:,:,0] = self.HPmontage*(defectMontage==0)*(maskMontage==0) + 255*(defectMontage==1)
            combined3D[:,:,1] = self.HPmontage*(defectMontage==0)*(maskMontage==0) + 128*(maskMontage==1)
            combined3D[:,:,2] = self.HPmontage*(defectMontage==0)*(maskMontage==0) + 255*(maskMontage==1)
        else:
            combined3D[:,:,0] = self.HPmontage*(defectMontage==0) + 255*(defectMontage==1)
            combined3D[:,:,1] = self.HPmontage*(defectMontage==0)
            combined3D[:,:,2] = self.HPmontage*(defectMontage==0)
        return combined3D.astype(np.uint8)
    
    def borderMontage(self):
        '''same as above but for the border array (not sure why I made the same function twice, but I did...)'''
        combined3D = np.zeros((self.HPmontage.shape[0],self.HPmontage.shape[1],3))
        defectMontage = self.montage(self.defectBorder)
        combined3D[:,:,0] = self.HPmontage*(defectMontage==0) + 255*(defectMontage==1)
        combined3D[:,:,1] = self.HPmontage*(defectMontage==0)
        combined3D[:,:,2] = self.HPmontage*(defectMontage==0)
        return combined3D.astype(np.uint8)

    def montage(self,arr):
        '''inputs 3D array [rows, cols, slices], returns 2D array montage [rows,cols*slices]'''
        l = [list(x) for x in arr.transpose(1,2,0).swapaxes(0,2)]
        montage = np.block(l)
        return montage

    def cropToData(self,A):
        '''Given a binary array (mask) will return the rows/columns/slices in which data exist.
            This allows us to crop out the empty rows/cols/slices for better display'''
        slices = np.multiply(np.sum(np.sum(A,axis=0),axis=0)>0,list(range(0,A.shape[2])))
        rows = np.multiply(np.sum(np.sum(A,axis=1),axis=1)>0,list(range(0,A.shape[0])))
        cols = np.multiply(np.sum(np.sum(A,axis=2),axis=0)>0,list(range(0,A.shape[1])))
        slices = [x for x in range(0,A.shape[2]) if slices[x]]
        rows = [x for x in range(0,A.shape[0]) if rows[x]]
        cols = [x for x in range(0,A.shape[1]) if cols[x]]
        # return A[np.ix_(rows,cols,slices)], rows, cols, slices
        return rows, cols, slices


## ---------------------------------------------- ## 
## -------------- Other helpers ----------------- ##
## ---------------------------------------------- ## 

def get_executable_directory():
    '''Returns the path of the executable, or the python script'''
    if getattr(sys, 'frozen', False):
        # Executable is frozen (compiled with PyInstaller)
        return os.path.dirname(sys.executable)
    else:
        # Executable is not frozen
        return os.path.dirname(os.path.realpath(__file__))

def load_Nifti_file(path):
    '''Opens a Nifti Dataset'''
    activeNifti  = nib.load(path)
    nii_data = activeNifti.get_fdata()
    nii_aff  = activeNifti.affine
    nii_hdr  = activeNifti.header
    return(nii_data,nii_aff,nii_hdr)

def drawArray(window,A,nCol,slider_value, sliceRange, whichImage = '-BORDERIMAGE-'):
    '''Updates the PySimpleGUI window with a new 2D array (grayscale) or 3D array (RGB 3D)'''
    nPixels = 200
    nRows = A.shape[0]
    imgAr = Image.fromarray(A[:,range(nCol*(slider_value-sliceRange),nCol*(slider_value+sliceRange))].astype(np.uint8))
    imgAr = imgAr.resize((int(nPixels*2*sliceRange*nCol/nRows),nPixels))
    image = ImageTk.PhotoImage(image=imgAr)
    window[whichImage].update(data=image)


def open_or_create_excel_file(parent_folder,XLname):
    '''Either opens an existing GUI results xlsx or creates one'''
    xlsx_filePath = os.path.join(parent_folder,XLname)
    try:
        workbook = openpyxl.load_workbook(xlsx_filePath)
        print(f"\033[32mOpened existing workbook:\033[37m {xlsx_filePath}")
        worksheet = workbook[workbook.sheetnames[0]]
        # -- What if the XL file exists, but is empty? Let's check that...
        if worksheet['A'][1].value == '':
            print('\033[31mWorksheet is empty. Gotta make a new one!\033[37m')
            raise Exception('Worksheet is empty. Gotta make a new one!')
    except:
        # -- Didn't find your file? That's ok, we'll just create a new one here
        workbook = openpyxl.Workbook()
        workbook.save(xlsx_filePath)
        print(f"\033[94mCreated new workbook:\033[37m {xlsx_filePath}")
        worksheet = workbook[workbook.sheetnames[0]]
        worksheet.cell(1,1,'File Name')
        worksheet.cell(1,2,'N4 corrected view')
        worksheet.cell(1,3,'Initial Threshold')
        worksheet.cell(1,4,'Chosen Threshold')
        worksheet.cell(1,5,'Calculated VDP')
        worksheet.cell(1,6,'VDP Estimate')
        worksheet.cell(1,7,'Quality Rank')
        worksheet.cell(1,8,'Disease Guess')
        worksheet.cell(1,9,'Disease Severity')
        worksheet.cell(1,10,'Artifacts')
        worksheet.cell(1,11,'Coil Shading')
        worksheet.cell(1,12,'Segmentation Errors')
        worksheet.cell(1,13,'Low SNR')
        worksheet.cell(1,14,'Slice Varying Threshold')
        worksheet.cell(1,15,'Vascular Defects')
        worksheet.cell(1,16,'Partial Voluming')
        worksheet.cell(1,17,'Time to Review')
        worksheet.cell(1,18,'Notes v240410_RPT')



        # -- We also need to create a list of cases to be reviewed
        # -- All data in the Nifti folder is used. Each case reviewed 3x
        fileList = os.listdir(os.path.join(parent_folder,'Niftis\\'))

        # -- Each case will be reviewed 4x - twice showing N4 corrected and twice showing raw HPG
        caseList =[]
        caseList.extend(fileList)
        caseList.extend(fileList)
        caseList.extend(fileList)
        caseList.extend(fileList)

        # -- Whether or not the case is viewed w/wo N4 is given by a binary array
        N4view = [np.zeros(len(fileList)*2),np.ones(len(fileList)*2)]
        N4view = np.concatenate(N4view)

        indices = np.arange(len(caseList))
        shuffle(indices)
        # -- Drop each file to be reviewed into the first column of the XL
        # -- N4 assignments go into the second column
        for k in range(len(indices)):
            worksheet.cell(k+2,1,caseList[indices[k]])
            worksheet.cell(k+2,2,N4view[indices[k]])

        workbook.save(xlsx_filePath)

    return workbook, xlsx_filePath

