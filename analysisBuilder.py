import numpy as np
from Vent_Analysis import Vent_Analysis
import nibabel as nii
import os
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans # ---------- for kMeans VDP
import openpyxl


workbook = openpyxl.Workbook()
workbook.save('c:/PIRL/data/VDP_GUI_thresholds.xlsx')
worksheet = workbook[workbook.sheetnames[0]]
worksheet.cell(1, 1, 'File')
worksheet.cell(1, 2, 'MA threshold')
worksheet.cell(1, 3, 'MA defect')
worksheet.cell(1, 4, 'LB threshold')
worksheet.cell(1, 5, 'LB defect')
worksheet.cell(1, 6, 'KM threshold')
worksheet.cell(1, 7, 'KM defect')

path = 'C:/VDP_GUI/Niftis/'
files = os.listdir(path)
k=2
for file in files:
    print(f"{k}/{len(files)} -- file")
    A = nii.loadsave.load(os.path.join(path,file)).get_fdata()
    xenon = A[:,:,:,0]
    mask = A[:,:,:,1]
    Vent1 = Vent_Analysis(xenon_array=A[:,:,:,0],mask_array=A[:,:,:,1])
    Vent1.vox = [1,1,1]
    Vent1.calculate_VDP()
    print(Vent1.defect_thresholds)
    worksheet.cell(k, 1, file)
    worksheet.cell(k, 2, float(Vent1.defect_thresholds[0]))
    worksheet.cell(k, 3, float(Vent1.metadata['VDP']))
    worksheet.cell(k, 4, float(Vent1.defect_thresholds[1]))
    worksheet.cell(k, 5, float(Vent1.metadata['VDP_lb']))
    worksheet.cell(k, 6, float(Vent1.defect_thresholds[2]))
    worksheet.cell(k, 7, float(Vent1.metadata['VDP_km']))
    k += 1



workbook.save('c:/PIRL/data/VDP_GUI_thresholds.xlsx')



makeSlide(Vent1.defectArrayKM[:,:,10:20])
# plt.imshow(clustered_xenon[:,:,15])
# plt.show()

def makeSlide(A):
    B = A[:,:,0]
    for k in np.arange(2,A.shape[2]):
        B = np.hstack((B,A[:,:,k]))
    plt.imshow(B)
    plt.show()

# makeSlide(xenon[:,:,10:20])
# makeSlide((xenon[:,:,10:20]<low_xenon_threshold)*mask[:,:,10:20])